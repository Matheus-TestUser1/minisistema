"""
Excel Reports Generator Module
Generates professional Excel reports using openpyxl
"""
import logging
from typing import Dict, Any, List
from datetime import datetime
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..database import ReportConfig

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """Generates Excel reports with professional formatting"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel report generation")
    
    def _create_workbook_with_style(self, title: str) -> Workbook:
        """Create a new workbook with default styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add generation info
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Add company info
        ws['A3'] = "🌲 MADEIREIRA MARIA LUIZA"
        ws['A3'].font = Font(size=12, bold=True)
        
        return wb
    
    def _style_headers(self, ws, start_row: int, end_col: int):
        """Apply header styling to a row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, end_col + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    def _style_data_row(self, ws, row: int, end_col: int, alternate: bool = False):
        """Apply data row styling"""
        fill_color = "F2F2F2" if alternate else "FFFFFF"
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_products_report(self, data: List[Dict[str, Any]], config: ReportConfig) -> Dict[str, Any]:
        """Generate products Excel report"""
        try:
            wb = self._create_workbook_with_style(config.titulo or "Relatório de Produtos")
            ws = wb.active
            
            if not data:
                ws['A5'] = "Nenhum produto encontrado"
                filename = self._save_workbook(wb, "produtos_vazio")
                return {'success': True, 'filename': filename, 'message': 'Relatório gerado (vazio)'}
            
            # Headers starting at row 5
            headers = list(data[0].keys())
            start_row = 5
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
            
            self._style_headers(ws, start_row, len(headers))
            
            # Data rows
            for row_idx, product in enumerate(data, start_row + 1):
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col, value=product[key])
                
                self._style_data_row(ws, row_idx, len(headers), row_idx % 2 == 0)
            
            # Summary
            summary_row = len(data) + start_row + 2
            ws.cell(row=summary_row, column=1, value=f"Total de produtos: {len(data)}")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            self._auto_adjust_columns(ws)
            filename = self._save_workbook(wb, "produtos")
            
            return {'success': True, 'filename': filename, 'total_records': len(data)}
            
        except Exception as e:
            logger.error(f"Error generating Excel products report: {e}")
            return {'success': False, 'error': str(e)}
    
    def generate_inventory_report(self, stock_data: Dict[str, Any], config: ReportConfig) -> Dict[str, Any]:
        """Generate inventory Excel report"""
        try:
            wb = self._create_workbook_with_style(config.titulo or "Relatório de Estoque")
            ws = wb.active
            
            # Summary section
            summary_start = 5
            ws.cell(row=summary_start, column=1, value="RESUMO DO ESTOQUE")
            ws.cell(row=summary_start, column=1).font = Font(size=14, bold=True)
            
            summary_data = [
                ("Total de Produtos:", stock_data.get('total_products', 0)),
                ("Valor Total em Estoque:", f"R$ {stock_data.get('total_stock_value', 0):.2f}"),
                ("Produtos em Falta:", stock_data.get('out_of_stock_count', 0)),
                ("Produtos com Estoque Baixo:", stock_data.get('low_stock_count', 0))
            ]
            
            for i, (label, value) in enumerate(summary_data, summary_start + 2):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
                ws.cell(row=i, column=1).font = Font(bold=True)
            
            # Categories section
            if stock_data.get('categories'):
                cat_start = summary_start + len(summary_data) + 3
                ws.cell(row=cat_start, column=1, value="RESUMO POR CATEGORIA")
                ws.cell(row=cat_start, column=1).font = Font(size=14, bold=True)
                
                cat_headers = ["Categoria", "Qtd Produtos", "Valor Total", "Estoque Total"]
                for col, header in enumerate(cat_headers, 1):
                    ws.cell(row=cat_start + 2, column=col, value=header)
                
                self._style_headers(ws, cat_start + 2, len(cat_headers))
                
                for row_idx, (categoria, data) in enumerate(stock_data['categories'].items(), cat_start + 3):
                    ws.cell(row=row_idx, column=1, value=categoria)
                    ws.cell(row=row_idx, column=2, value=data['count'])
                    ws.cell(row=row_idx, column=3, value=f"R$ {data['total_value']:.2f}")
                    ws.cell(row=row_idx, column=4, value=data['total_stock'])
                    
                    self._style_data_row(ws, row_idx, len(cat_headers), row_idx % 2 == 0)
            
            # Products detail
            if stock_data.get('products'):
                products = stock_data['products']
                detail_start = cat_start + len(stock_data.get('categories', {})) + 5 if stock_data.get('categories') else summary_start + len(summary_data) + 5
                
                ws.cell(row=detail_start, column=1, value="DETALHAMENTO DOS PRODUTOS")
                ws.cell(row=detail_start, column=1).font = Font(size=14, bold=True)
                
                headers = ["Código", "Descrição", "Categoria", "Estoque", "Preço", "Valor Estoque", "Status"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=detail_start + 2, column=col, value=header)
                
                self._style_headers(ws, detail_start + 2, len(headers))
                
                for row_idx, product in enumerate(products, detail_start + 3):
                    row_data = [
                        product['codigo'],
                        product['descricao'],
                        product['categoria'],
                        product['estoque_atual'],
                        f"R$ {product['preco_venda']:.2f}",
                        f"R$ {product['valor_estoque']:.2f}",
                        product['status']
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col, value=value)
                    
                    self._style_data_row(ws, row_idx, len(headers), row_idx % 2 == 0)
            
            self._auto_adjust_columns(ws)
            filename = self._save_workbook(wb, "estoque")
            
            return {'success': True, 'filename': filename}
            
        except Exception as e:
            logger.error(f"Error generating Excel inventory report: {e}")
            return {'success': False, 'error': str(e)}
    
    def generate_low_stock_report(self, alerts: List[Dict[str, Any]], config: ReportConfig) -> Dict[str, Any]:
        """Generate low stock Excel report"""
        try:
            wb = self._create_workbook_with_style(config.titulo or "Produtos em Falta")
            ws = wb.active
            
            if not alerts:
                ws['A5'] = "Nenhum produto com estoque baixo encontrado"
                filename = self._save_workbook(wb, "estoque_baixo_vazio")
                return {'success': True, 'filename': filename, 'message': 'Nenhum produto em falta'}
            
            # Headers
            headers = ["Código", "Descrição", "Categoria", "Estoque Atual", "Urgência"]
            start_row = 5
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
            
            self._style_headers(ws, start_row, len(headers))
            
            # Data
            for row_idx, alert in enumerate(alerts, start_row + 1):
                row_data = [
                    alert['codigo'],
                    alert['descricao'],
                    alert['categoria'] or 'Sem Categoria',
                    alert['estoque_atual'],
                    alert['urgency']
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    
                    # Color code by urgency
                    if alert['urgency'] == 'CRITICAL':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif alert['urgency'] == 'HIGH':
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                
                self._style_data_row(ws, row_idx, len(headers), False)
            
            self._auto_adjust_columns(ws)
            filename = self._save_workbook(wb, "estoque_baixo")
            
            return {'success': True, 'filename': filename, 'total_alerts': len(alerts)}
            
        except Exception as e:
            logger.error(f"Error generating Excel low stock report: {e}")
            return {'success': False, 'error': str(e)}
    
    def generate_price_list(self, price_data: List[Dict[str, Any]], config: ReportConfig) -> Dict[str, Any]:
        """Generate price list Excel report"""
        try:
            wb = self._create_workbook_with_style(config.titulo or "Lista de Preços")
            ws = wb.active
            
            # Headers
            headers = ["Código", "Descrição", "Categoria", "Preço de Venda", "Unidade"]
            start_row = 5
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
            
            self._style_headers(ws, start_row, len(headers))
            
            # Data grouped by category
            current_category = None
            row_idx = start_row + 1
            
            for item in price_data:
                if item['Categoria'] != current_category:
                    if current_category is not None:
                        row_idx += 1  # Add space between categories
                    
                    # Category header
                    ws.cell(row=row_idx, column=1, value=f"CATEGORIA: {item['Categoria']}")
                    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    current_category = item['Categoria']
                    row_idx += 1
                
                # Product data
                row_data = [
                    item['Código'],
                    item['Descrição'],
                    item['Categoria'],
                    item['Preço de Venda'],
                    item['Unidade']
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
                
                self._style_data_row(ws, row_idx, len(headers), row_idx % 2 == 0)
                row_idx += 1
            
            self._auto_adjust_columns(ws)
            filename = self._save_workbook(wb, "lista_precos")
            
            return {'success': True, 'filename': filename, 'total_items': len(price_data)}
            
        except Exception as e:
            logger.error(f"Error generating Excel price list: {e}")
            return {'success': False, 'error': str(e)}
    
    def generate_category_report(self, categories: Dict[str, Dict], config: ReportConfig) -> Dict[str, Any]:
        """Generate category-based Excel report"""
        try:
            wb = self._create_workbook_with_style(config.titulo or "Relatório por Categoria")
            
            for categoria, data in categories.items():
                # Create a sheet for each category
                ws = wb.create_sheet(title=categoria[:30])  # Sheet names limited to 30 chars
                
                # Category summary
                ws['A1'] = f"CATEGORIA: {categoria}"
                ws['A1'].font = Font(size=16, bold=True)
                
                ws['A3'] = f"Total de Produtos: {data['total_items']}"
                ws['A4'] = f"Valor Total: R$ {data['valor_total']:.2f}"
                ws['A5'] = f"Estoque Total: {data['estoque_total']} unidades"
                
                # Products in category
                headers = ["Código", "Descrição", "Preço", "Estoque", "Valor Estoque"]
                start_row = 7
                
                for col, header in enumerate(headers, 1):
                    ws.cell(row=start_row, column=col, value=header)
                
                self._style_headers(ws, start_row, len(headers))
                
                for row_idx, product in enumerate(data['produtos'], start_row + 1):
                    row_data = [
                        product.codigo,
                        product.descricao,
                        f"R$ {product.preco_venda:.2f}",
                        product.estoque_atual,
                        f"R$ {(product.preco_venda * product.estoque_atual):.2f}"
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col, value=value)
                    
                    self._style_data_row(ws, row_idx, len(headers), row_idx % 2 == 0)
                
                self._auto_adjust_columns(ws)
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            filename = self._save_workbook(wb, "categorias")
            
            return {'success': True, 'filename': filename, 'categories_count': len(categories)}
            
        except Exception as e:
            logger.error(f"Error generating Excel category report: {e}")
            return {'success': False, 'error': str(e)}
    
    def generate_reorder_report(self, suggestions: List[Dict[str, Any]], config: ReportConfig) -> Dict[str, Any]:
        """Generate reorder suggestions Excel report"""
        try:
            wb = self._create_workbook_with_style(config.titulo or "Sugestões de Reposição")
            ws = wb.active
            
            if not suggestions:
                ws['A5'] = "Nenhuma sugestão de reposição no momento"
                filename = self._save_workbook(wb, "reposicao_vazio")
                return {'success': True, 'filename': filename}
            
            # Headers
            headers = ["Código", "Descrição", "Categoria", "Estoque Atual", "Qtd Sugerida", "Prioridade"]
            start_row = 5
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
            
            self._style_headers(ws, start_row, len(headers))
            
            # Data
            for row_idx, suggestion in enumerate(suggestions, start_row + 1):
                row_data = [
                    suggestion['codigo'],
                    suggestion['descricao'],
                    suggestion['categoria'] or 'Sem Categoria',
                    suggestion['estoque_atual'],
                    suggestion['quantidade_sugerida'],
                    suggestion['priority']
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    
                    # Color code by priority
                    if suggestion['priority'] == 'HIGH':
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                
                self._style_data_row(ws, row_idx, len(headers), False)
            
            self._auto_adjust_columns(ws)
            filename = self._save_workbook(wb, "reposicao")
            
            return {'success': True, 'filename': filename, 'suggestions_count': len(suggestions)}
            
        except Exception as e:
            logger.error(f"Error generating Excel reorder report: {e}")
            return {'success': False, 'error': str(e)}
    
    def _save_workbook(self, wb: Workbook, prefix: str) -> str:
        """Save workbook and return filename"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{prefix}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")
        return filename